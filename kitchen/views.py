from django.contrib.auth.decorators import login_required, user_passes_test
from django.shortcuts import render, get_object_or_404, redirect
from django.urls import reverse

from django.views.generic import ListView, DetailView
from django.contrib.auth.mixins import LoginRequiredMixin, UserPassesTestMixin
from django.db.models import Q
from django.views.decorators.http import require_POST
from django.contrib import messages

def is_kitchen_admin(user):
    return user.is_authenticated and (user.is_staff or user.is_superuser)

import os
import requests
import csv
import io
import datetime
from django.utils.dateparse import parse_date
from django.utils import timezone
from django.db.models import Sum, Count, Q
from django.contrib.auth import get_user_model
from vending.models import Order, OrderStatus, OrderItem, PlanType, PlanSubType
from ai_agents.models import AgentActivity, AgentInteractionLog
from django.http import JsonResponse
from django.views.generic import TemplateView
import concurrent.futures

User = get_user_model()

class DashboardView(LoginRequiredMixin, UserPassesTestMixin, ListView):
    model = Order
    template_name = 'kitchen/dashboard.html'
    context_object_name = 'orders'

    def test_func(self):
        return is_kitchen_admin(self.request.user)

    def get_queryset(self):
        return Order.objects.filter(
            status__in=[
                OrderStatus.PENDING,
                OrderStatus.PREPARING,
                OrderStatus.READY,
                OrderStatus.CONFIRMED,
                OrderStatus.PENDING_FULFILLMENT,
            ]
        ).distinct().order_by('-created_at')

    def get_context_data(self, **kwargs):
        context = super().get_context_data(**kwargs)
        from vending.models import VendingLocation, MasterItem
        from catering.models import SweetsItem
        context['pending_fulfillment_orders'] = Order.objects.filter(
            status=OrderStatus.PENDING_FULFILLMENT
        ).order_by('-created_at')
        context['locations'] = VendingLocation.objects.all().order_by('name')

        master_items = list(MasterItem.objects.values_list('name', flat=True))
        sweets_items = list(SweetsItem.objects.values_list('name', flat=True))
        context['item_names'] = sorted(list(set(master_items + sweets_items)))

        # Beit Nahla orders are rendered in their own tab on this dashboard
        # (the user prefers a single Active Orders screen over a separate
        # sidebar entry).
        bn_orders = (
            BeitNahlaOrder.objects
            .prefetch_related('items')
            .order_by('-created_at')[:200]
        )
        context['bn_orders_json'] = [_bn_order_payload(o) for o in bn_orders]
        context['bn_status_choices'] = [
            {'value': v, 'label': l} for v, l in BeitNahlaOrderStatus.choices
        ]
        return context

class TrackingView(LoginRequiredMixin, UserPassesTestMixin, ListView):
    model = Order
    template_name = 'kitchen/tracking.html'
    context_object_name = 'orders'
    
    def test_func(self):
        return is_kitchen_admin(self.request.user)
    ordering = ['-created_at']

    def get_queryset(self):
        from django.db.models import Q
        # Tracking needs ALL orders
        qs = Order.objects.all().order_by('-created_at')
        
        status_filter = self.request.GET.get('status')
        if status_filter:
            qs = qs.filter(status=status_filter)
            
        location_filters = self.request.GET.getlist('location')
        if location_filters and any(location_filters):
            # filter out empty strings
            valid_locs = [loc for loc in location_filters if loc]
            if valid_locs:
                qs = qs.filter(location_id__in=valid_locs)
            
        item_filters = self.request.GET.getlist('item_search')
        if item_filters and any(item_filters):
            valid_items = [item for item in item_filters if item]
            if valid_items:
                item_q = Q()
                for item_filter in valid_items:
                    item_q |= Q(items__item_name_snapshot__icontains=item_filter)
                    item_q |= Q(items__menu_item__name__icontains=item_filter)
                    item_q |= Q(items__master_item__name__icontains=item_filter)
                    item_q |= Q(items__sweets_item__name__icontains=item_filter)
                qs = qs.filter(item_q).distinct()
            
        return qs

    def get_context_data(self, **kwargs):
        context = super().get_context_data(**kwargs)
        from vending.models import VendingLocation, MasterItem
        from catering.models import SweetsItem
        context['status_choices'] = OrderStatus.choices
        context['locations'] = VendingLocation.objects.all().order_by('name')
        
        master_items = list(MasterItem.objects.values_list('name', flat=True))
        sweets_items = list(SweetsItem.objects.values_list('name', flat=True))
        context['item_names'] = sorted(list(set(master_items + sweets_items)))
        
        # Pass selected filters back to template
        context['selected_locations'] = self.request.GET.getlist('location')
        context['selected_items'] = self.request.GET.getlist('item_search')
        
        return context

class AnalyticsDashboardView(LoginRequiredMixin, UserPassesTestMixin, TemplateView):
    template_name = 'kitchen/analytics.html'

    def test_func(self):
        return is_kitchen_admin(self.request.user)

    def get_context_data(self, **kwargs):
        context = super().get_context_data(**kwargs)
        
        # 1. Date Filtering
        start_date_str = self.request.GET.get('start_date')
        end_date_str = self.request.GET.get('end_date')
        
        today = timezone.localtime().date()
        
        if start_date_str:
            start_date = parse_date(start_date_str)
        else:
            # Default to last 30 days if no filter
            start_date = today - datetime.timedelta(days=30)
            
        if end_date_str:
            end_date = parse_date(end_date_str)
        else:
            end_date = today
            
        # Ensure we cover the full end day
        end_datetime = datetime.datetime.combine(end_date, datetime.time.max)
        if timezone.is_naive(end_datetime):
            end_datetime = timezone.make_aware(end_datetime)
            
        start_datetime = datetime.datetime.combine(start_date, datetime.time.min)
        if timezone.is_naive(start_datetime):
             start_datetime = timezone.make_aware(start_datetime)

        # 2. Users Stats
        users_in_range_qs = User.objects.filter(date_joined__range=(start_datetime, end_datetime))
        total_users = User.objects.count()
        users_in_range = users_in_range_qs.count()

        # 3. Orders & Revenue Stats
        # Today's orders
        today_start = timezone.make_aware(datetime.datetime.combine(today, datetime.time.min))
        today_orders = Order.objects.filter(created_at__gte=today_start).exclude(status=OrderStatus.DRAFT)
        today_orders_count = today_orders.count()
        today_revenue = today_orders.aggregate(total=Sum('total_amount'))['total'] or 0.00
        
        # Range orders
        range_orders = Order.objects.filter(
            created_at__range=(start_datetime, end_datetime)
        ).exclude(status=OrderStatus.DRAFT)
        range_orders_count = range_orders.count()
        range_revenue = range_orders.aggregate(total=Sum('total_amount'))['total'] or 0.00

        # 4. User Leaderboard (Only users joined in range)
        user_stats = User.objects.filter(
            date_joined__range=(start_datetime, end_datetime)
        ).annotate(
            orders_count=Count('orders', filter=Q(orders__created_at__range=(start_datetime, end_datetime), orders__status__exact=OrderStatus.DRAFT, _negated=True)),
            total_spent=Sum('orders__total_amount', filter=Q(orders__created_at__range=(start_datetime, end_datetime), orders__status__exact=OrderStatus.DRAFT, _negated=True))
        ).order_by('-date_joined')

        context.update({
            'start_date': start_date.strftime('%Y-%m-%d'),
            'end_date': end_date.strftime('%Y-%m-%d'),
            'total_users': total_users,
            'users_in_range': users_in_range,
            'today_orders_count': today_orders_count,
            'today_revenue': today_revenue,
            'range_orders_count': range_orders_count,
            'range_revenue': range_revenue,
            'user_stats': user_stats,
        })
        return context

class AccountsDashboardView(LoginRequiredMixin, UserPassesTestMixin, TemplateView):
    template_name = 'kitchen/accounts.html'

    def test_func(self):
        return is_kitchen_admin(self.request.user)

    def get_context_data(self, **kwargs):
        context = super().get_context_data(**kwargs)
        from django.db.models.functions import TruncDate
        
        # 1. Date Filtering
        start_date_str = self.request.GET.get('start_date')
        end_date_str = self.request.GET.get('end_date')
        period = self.request.GET.get('period', 'custom')
        
        today = timezone.localtime().date()
        
        if period == 'weekly':
            start_date = today - datetime.timedelta(days=7)
            end_date = today
        elif period == 'monthly':
            start_date = today - datetime.timedelta(days=30)
            end_date = today
        else:
            if start_date_str:
                start_date = parse_date(start_date_str)
            else:
                start_date = today - datetime.timedelta(days=30)
            if end_date_str:
                end_date = parse_date(end_date_str)
            else:
                end_date = today
            
        # Ensure we cover the full end day
        end_datetime = datetime.datetime.combine(end_date, datetime.time.max)
        if timezone.is_naive(end_datetime):
            end_datetime = timezone.make_aware(end_datetime)
            
        start_datetime = datetime.datetime.combine(start_date, datetime.time.min)
        if timezone.is_naive(start_datetime):
             start_datetime = timezone.make_aware(start_datetime)

        # 2. Filtered Orders
        base_qs = Order.objects.filter(
            created_at__range=(start_datetime, end_datetime)
        ).exclude(status=OrderStatus.DRAFT)

        total_orders = base_qs.count()
        completed_orders = base_qs.filter(status=OrderStatus.COMPLETED).count()
        ready_orders = base_qs.filter(status=OrderStatus.READY).count()
        preparing_orders = base_qs.filter(status__in=[OrderStatus.PREPARING, OrderStatus.CONFIRMED]).count()
        pending_orders = base_qs.filter(status=OrderStatus.PENDING).count()
        
        total_earnings = base_qs.aggregate(total=Sum('total_amount'))['total'] or 0.00
        
        # 3. Daily Stats for Chart (Fill missing dates with 0)
        query_stats = base_qs.annotate(date=TruncDate('created_at')) \
            .values('date') \
            .annotate(
                count=Count('id'),
                earnings=Sum('total_amount')
            ).order_by('date')
        
        stats_map = {s['date']: s for s in query_stats}
        full_daily_stats = []
        curr = start_date
        while curr <= end_date:
            if curr in stats_map:
                full_daily_stats.append(stats_map[curr])
            else:
                full_daily_stats.append({
                    'date': curr,
                    'count': 0,
                    'earnings': 0
                })
            curr += datetime.timedelta(days=1)

        # 4. Location Stats
        location_stats = base_qs.values('location__name') \
            .annotate(
                order_count=Count('id'),
                earnings=Sum('total_amount')
            ).order_by('-earnings')

        # 5. Pre-calculate rates for template
        avg_ticket = float(total_earnings) / total_orders if total_orders > 0 else 0
        completion_rate = (completed_orders / total_orders * 100) if total_orders > 0 else 0

        context.update({
            'start_date': start_date.strftime('%Y-%m-%d'),
            'end_date': end_date.strftime('%Y-%m-%d'),
            'period': period,
            'total_orders': total_orders,
            'completed_orders': completed_orders,
            'ready_orders': ready_orders,
            'preparing_orders': preparing_orders,
            'pending_orders': pending_orders,
            'total_earnings': total_earnings,
            'avg_ticket': avg_ticket,
            'completion_rate': completion_rate,
            'daily_stats': full_daily_stats,
            'location_stats': location_stats,
        })
        return context

class OrderDetailView(LoginRequiredMixin, UserPassesTestMixin, DetailView):
    model = Order
    template_name = 'kitchen/order_detail.html'
    context_object_name = 'order'

    def test_func(self):
        return is_kitchen_admin(self.request.user)

@login_required
@user_passes_test(is_kitchen_admin)
@require_POST
def update_order_status(request, pk):
    order = get_object_or_404(Order, pk=pk)
    new_status = request.POST.get('status')
    is_ajax = request.headers.get('X-Requested-With') == 'XMLHttpRequest'

    if new_status in OrderStatus.values:
        order.status = new_status
        order.save()
        if is_ajax:
            return JsonResponse({'success': True, 'new_status': new_status, 'display': order.get_status_display()})

    if is_ajax:
        return JsonResponse({'success': False, 'error': 'Invalid status'}, status=400)

    # Non-AJAX: redirect back to wherever the user came from
    referer = request.META.get('HTTP_REFERER', '')
    from django.urls import reverse
    if referer:
        return redirect(referer)
    return redirect('kitchen:tracking_dashboard')


# -----------------------------------------------------------
# ITEM STATUS UPDATE (Daily Orders)
# -----------------------------------------------------------
@login_required
@user_passes_test(is_kitchen_admin)
@require_POST
def update_item_status(request, pk):
    item = get_object_or_404(OrderItem, pk=pk)
    # Toggle 'READY' or 'PENDING'
    # Check current status found in item context or just use a query param
    # For now, let's assume we toggle between PENDING and READY
    
    current = item.status
    if current == OrderStatus.READY:
        item.status = OrderStatus.PENDING
    else:
        item.status = OrderStatus.READY
    
    item.save()
    
    # Check if all items in order are ready, then update order status?
    # Optional logic:
    # if not item.order.items.exclude(status=OrderStatus.READY).exists():
    #     item.order.status = OrderStatus.READY
    #     item.order.save()

    # Redirect back to daily orders with same date
    date_str = request.POST.get('date_str')
    url = reverse('kitchen:daily_orders')
    if date_str:
        url += f'?date={date_str}'
    return redirect(url)

# -----------------------------------------------------------
# MENU UPLOAD
# -----------------------------------------------------------
# -----------------------------------------------------------
# MENU UPLOAD
# -----------------------------------------------------------
from django.http import JsonResponse
import csv
import io
import re
from vending.models import Menu, MenuItem, DayOfWeek, VendingMachineStock
import requests
from django.core.files.base import ContentFile
from django.utils.text import slugify
import logging

logger = logging.getLogger(__name__)

# @login_required
# @user_passes_test(is_kitchen_admin)
@login_required
@user_passes_test(is_kitchen_admin)
def menu_upload_view(request):
    if request.method == 'POST' and request.FILES.get('menu_file'):
        # Check confirmation checkbox
        if request.POST.get('confirm_delete') != 'on':
            messages.error(request, "You must check the confirmation box to proceed with the upload.")
            return render(request, 'kitchen/menu_upload.html')

        file = request.FILES['menu_file']
        
        try:
            data = []
            if file.name.endswith('.csv'):
                decoded_file = file.read().decode('utf-8-sig').splitlines()
                reader = csv.DictReader(decoded_file)
                data = list(reader)
            elif file.name.endswith(('.xls', '.xlsx')):
                import openpyxl
                wb = openpyxl.load_workbook(file)
                sheet = wb.active
                
                rows = list(sheet.iter_rows(values_only=True))
                if not rows:
                    raise ValueError("Empty file")
                    
                headers = [str(h).strip() if h else f"col_{i}" for i, h in enumerate(rows[0])]
                for row in rows[1:]:
                    if not any(row):
                        continue
                    row_dict = {}
                    for i, val in enumerate(row):
                        if i < len(headers):
                            row_dict[headers[i]] = val
                    data.append(row_dict)
            else:
                messages.error(request, "Unsupported file format. Please use CSV or Excel.")
                return render(request, 'kitchen/menu_upload.html')

            # DESTRUCTIVE WIPE & IMPORT
            from django.db import transaction
            from vending.models import Menu, MenuItem, MasterItem
            
            with transaction.atomic():
                MenuItem.objects.all().delete()
                Menu.objects.all().delete()
                MasterItem.objects.all().delete()
                
                masters_created, schedules_created, logs, pending_images = import_vending_sheet(data)
                
            # PERFORM IMAGE DOWNLOADS OUTSIDE TRANSACTION TO PREVENT DB LOCKS
            import time
            import threading
            from django.core.files.base import ContentFile
            from django.utils.text import slugify
            
            def download_images_task(images):
                from vending.models import MasterItem
                for img_info in images:
                    master_id = img_info['master_id']
                    item_name = img_info['item_name']
                    picture_url = img_info['picture_url']
                    field_name = img_info.get('field_name', 'image') # Default to primary image
                    
                    try:
                        direct_url = get_google_drive_direct_link(picture_url)
                        max_retries = 3
                        
                        for attempt in range(max_retries):
                            try:
                                response = requests.get(direct_url, timeout=30)
                                if response.status_code == 200:
                                    filename = f"{slugify(item_name)}_{field_name}.jpg"
                                    master = MasterItem.objects.get(id=master_id)
                                    getattr(master, field_name).save(filename, ContentFile(response.content), save=True)
                                    break # Success
                                elif response.status_code == 429:
                                    if attempt == max_retries - 1:
                                        logs.append(f"Image download failed for {item_name}: Rate limited (HTTP 429)")
                                    else:
                                        time.sleep(2 ** attempt)
                                else:
                                    if attempt == max_retries - 1:
                                        logs.append(f"Image download failed for {item_name}: HTTP {response.status_code}")
                                    else:
                                        time.sleep(2)
                            except Exception as req_e:
                                if attempt == max_retries - 1:
                                    logs.append(f"Image download failed for {item_name}: {req_e}")
                                else:
                                    time.sleep(2 ** attempt)
                    except Exception as e:
                        logs.append(f"Image error for {item_name}: {e}")

            if pending_images:
                thread = threading.Thread(target=download_images_task, args=(pending_images,))
                thread.start()
                messages.info(request, f"Started downloading {len(pending_images)} new images in the background.")
                
            messages.success(request, f"Successfully imported {masters_created} items and linked {schedules_created} schedules.")
            # We can log warnings to messages as well if desired
            for log in logs[:5]: # Show max 5 warnings
                messages.warning(request, log)

        except Exception as e:
            logger.error(f"Error processing file: {e}")
            messages.error(request, f"Error processing file: {str(e)}")
            
    return render(request, 'kitchen/menu_upload.html')

def parse_numeric(val_str):
    """Safely parse numbers from strings like '35 AED' or '36g'.
    Handles commas by treating them as decimal separators if only one exists,
    otherwise removes them as thousand separators.
    """
    if not val_str:
        return 0
    
    val_str = str(val_str).strip()
    
    # If there is a comma but no dot, it's likely a European-style decimal separator
    if ',' in val_str and '.' not in val_str:
        # Check if it looks like a decimal separator (at the end)
        parts = val_str.split(',')
        if len(parts) == 2 and len(parts[1]) <= 2:
            val_str = val_str.replace(',', '.')
            
    # Remove everything except digits and dots
    cleaned = re.sub(r'[^\d.]', '', val_str)
    
    # If multiple dots, keep only the first one
    if cleaned.count('.') > 1:
        parts = cleaned.split('.')
        cleaned = parts[0] + '.' + ''.join(parts[1:])
        
    try:
        val = float(cleaned) if cleaned else 0
        # Cap at 1 million to prevent crazy DB overflows
        if val > 1000000:
            return 1000000
        return val
    except ValueError:
        return 0

def get_google_drive_direct_link(url):
    """
    Transforms a Google Drive sharing link into a direct download link.
    If the URL is not a recognized Google Drive link, returns the original URL.
    """
    if not url or "drive.google.com" not in url:
        return url
        
    import re
    # Match /file/d/ID/...
    match = re.search(r'/file/d/([a-zA-Z0-9_-]+)', url)
    if match:
        file_id = match.group(1)
        return f"https://drive.google.com/uc?export=download&id={file_id}"
        
    # Match id=ID
    match = re.search(r'id=([a-zA-Z0-9_-]+)', url)
    if match:
        file_id = match.group(1)
        return f"https://drive.google.com/uc?export=download&id={file_id}"
        
    return url

def import_vending_sheet(data_iter):
    from vending.models import Menu, MenuItem, MasterItem, DayOfWeek
    import requests
    from django.core.files.base import ContentFile
    from django.utils.text import slugify

    day_map = {
        'Monday': DayOfWeek.MONDAY,
        'Tuesday': DayOfWeek.TUESDAY, 
        'Wednesday': DayOfWeek.WEDNESDAY,
        'Thursday': DayOfWeek.THURSDAY,
        'Friday': DayOfWeek.FRIDAY,
        'Saturday': DayOfWeek.SATURDAY,
        'Sunday': DayOfWeek.SUNDAY
    }

    masters_count = 0
    schedules_count = 0
    logs = []
    pending_images = []

    for row in data_iter:
        row = {k.strip(): str(v).strip() for k, v in row.items() if k}
        
        # Parse Required Fields
        item_name = row.get('Item', row.get('Item Name', '')).strip()
        if not item_name:
            continue  # Skip empty names
            
        desc = row.get('Description', row.get('Item Description', ''))
        price_val = parse_numeric(row.get('Price', row.get('Item Price', 0)))
        
        # Parse Image URLs
        picture_url = row.get('Image', row.get('Picture', '')).strip()
        if picture_url.startswith('=') or 'IMAGE(' in picture_url:
            url_match = re.search(r'(https?://[^\s"\'\)]+)', picture_url)
            if url_match:
                picture_url = url_match.group(1)
        picture_url = picture_url.strip('"\'')

        picture_url2 = row.get('Image 2', row.get('Image2', row.get('Picture 2', ''))).strip()
        if picture_url2.startswith('=') or 'IMAGE(' in picture_url2:
            url_match2 = re.search(r'(https?://[^\s"\'\)]+)', picture_url2)
            if url_match2:
                picture_url2 = url_match2.group(1)
        picture_url2 = picture_url2.strip('"\'')

        # Create MasterItem ALWAYS
        master, created = MasterItem.objects.get_or_create(
            name=item_name,
            defaults={
                'description': desc,
                'default_price': price_val,
                'image_source_url': picture_url if picture_url.startswith('http') else None,
                'image2_source_url': picture_url2 if picture_url2.startswith('http') else None
            }
        )
        if created:
            masters_count += 1
            # Queue image for download after transaction if it's new
            if picture_url and picture_url.startswith('http'):
                pending_images.append({
                    'master_id': master.id,
                    'item_name': item_name,
                    'picture_url': picture_url,
                    'field_name': 'image'
                })
            if picture_url2 and picture_url2.startswith('http'):
                pending_images.append({
                    'master_id': master.id,
                    'item_name': item_name,
                    'picture_url': picture_url2,
                    'field_name': 'image2'
                })
        else:
            logs.append(f"Skipped duplicate MasterItem name: {item_name}")
            continue # We already processed this master (e.g. from duplicate rows in sheet)

        # Handle Optional Schedule Data
        week_raw = row.get('Week', '').strip().lower()
        day_raw = row.get('Day', '').strip().capitalize()
        
        if week_raw and day_raw in day_map:
            week_match = re.search(r'\d+', week_raw)
            week_num = int(week_match.group()) if week_match else 1
            
            menu, _ = Menu.objects.get_or_create(
                day_of_week=day_map[day_raw],
                week_number=week_num
            )
            
            MenuItem.objects.create(
                menu=menu,
                master_item=master,
                name=master.name,
                price=master.default_price,
                description=master.description,
                image=master.image,
                image_source_url=master.image_source_url
            )
            schedules_count += 1

    return masters_count, schedules_count, logs, pending_images

# -----------------------------------------------------------
# VENDING MASTER ITEM EDIT APIs
# -----------------------------------------------------------
from django.views.decorators.http import require_POST
import json

@login_required
@user_passes_test(is_kitchen_admin)
@require_POST
def vending_master_item_create_view(request):
    """Creates a new MasterItem from the Add Item modal."""
    from vending.models import MasterItem

    name = request.POST.get('name', '').strip()
    if not name:
        messages.error(request, "Item name is required.")
        return redirect('kitchen:vending_master_list')

    master = MasterItem(
        name=name,
        description=request.POST.get('description', ''),
        default_price=parse_numeric(request.POST.get('price', 0)),
        heating=request.POST.get('heating') == 'true',
        maximum_heating=int(request.POST.get('maximum_heating', 0) or 0),
    )

    if 'image' in request.FILES:
        master.image = request.FILES['image']
    if 'image2' in request.FILES:
        master.image2 = request.FILES['image2']

    try:
        master.save()
        messages.success(request, f"Master item '{master.name}' created successfully.")
    except Exception as e:
        from django.db import IntegrityError
        if 'unique' in str(e).lower() or 'UNIQUE' in str(e):
            messages.error(request, f"A master item named '{name}' already exists. Please choose a different name.")
        else:
            messages.error(request, f"Could not create item: {e}")
    return redirect('kitchen:vending_master_list')


@login_required
@user_passes_test(is_kitchen_admin)
@require_POST
def vending_master_item_delete_view(request, pk):
    """Deletes a MasterItem and all its associated MenuItems."""
    from vending.models import MasterItem
    master = get_object_or_404(MasterItem, pk=pk)
    name = master.name
    try:
        master.delete()
        messages.success(request, f"Master item '{name}' deleted successfully.")
    except Exception as e:
        messages.error(request, f"Could not delete '{name}': {e}")
    return redirect('kitchen:vending_master_list')


@login_required
@user_passes_test(is_kitchen_admin)
@require_POST
def vending_master_item_edit_view(request, pk):
    """Handles the edit modal form submission"""
    from vending.models import MasterItem, Menu, MenuItem, DayOfWeek
    
    master = get_object_or_404(MasterItem, pk=pk)
    
    # 1. Update Master Fields
    master.name = request.POST.get('name', master.name)
    master.description = request.POST.get('description', '')
    master.default_price = parse_numeric(request.POST.get('price', master.default_price))
    master.heating = request.POST.get('heating') == 'true'
    master.maximum_heating = int(request.POST.get('maximum_heating', 0) or 0)

    if 'image' in request.FILES:
        master.image = request.FILES['image']

    if 'image2' in request.FILES:
        master.image2 = request.FILES['image2']

    master.save() # Triggers signal to update any existing menu items' name/desc
    
    # 2. Update Schedules (Diff-based to preserve Cart links)
    schedules_json = request.POST.get('schedules_json', '[]')
    try:
        new_schedules = json.loads(schedules_json)
        day_map = {
            'Monday': DayOfWeek.MONDAY, 'Tuesday': DayOfWeek.TUESDAY, 
            'Wednesday': DayOfWeek.WEDNESDAY, 'Thursday': DayOfWeek.THURSDAY, 
            'Friday': DayOfWeek.FRIDAY, 'Saturday': DayOfWeek.SATURDAY, 
            'Sunday': DayOfWeek.SUNDAY
        }
        
        # Track intended schedule (week, day)
        intended_schedule = set()
        for s in new_schedules:
            week_num = int(s.get('week', 1))
            day_raw = s.get('day', 'Monday')
            day_val = day_map.get(day_raw, DayOfWeek.MONDAY)
            intended_schedule.add((week_num, day_val))
            
        # 1. Delete MenuItem records that are no longer in the schedule
        existing_items = MenuItem.objects.filter(master_item=master, menu__menu_type=MenuType.MONTHLY)
        for item in existing_items:
            if (item.menu.week_number, item.menu.day_of_week) not in intended_schedule:
                item.delete()
        
        # 2. Create or Update intended MenuItem records
        for week_num, day_val in intended_schedule:
            menu, _ = Menu.objects.get_or_create(
                day_of_week=day_val,
                week_number=week_num,
                menu_type=MenuType.MONTHLY
            )
            
            # Update existing or Create new
            MenuItem.objects.update_or_create(
                menu=menu,
                master_item=master,
                defaults={
                    'name': master.name,
                    'price': master.default_price,
                    'description': master.description,
                    'image': master.image,
                    'image_source_url': master.image_source_url
                }
            )
            
    except Exception as e:
        messages.error(request, f"Error updating schedules: {str(e)}")
        
    messages.success(request, f"Updated '{master.name}' and its schedules successfully.")
    return redirect('kitchen:vending_master_list')

@login_required
@user_passes_test(is_kitchen_admin)
def vending_master_schedule_api(request, pk):
    """Returns JSON of current schedules for a master item to populate the modal"""
    from vending.models import MasterItem, MenuType
    master = get_object_or_404(MasterItem, pk=pk)
    
    # Monthly schedules are the main editable list in the modal
    schedules = []
    for menu_item in master.menu_items.filter(menu__menu_type=MenuType.MONTHLY).select_related('menu'):
        schedules.append({
            'week': menu_item.menu.week_number,
            'day': menu_item.menu.day_of_week
        })
    
    # Weekly schedules shown read-only in modal
    weekly_schedules = []
    for menu_item in master.menu_items.filter(menu__menu_type=MenuType.WEEKLY).select_related('menu'):
        weekly_schedules.append({
            'day': menu_item.menu.day_of_week
        })
        
    return JsonResponse({
        'schedules': schedules,
        'weekly_schedules': weekly_schedules
    })


@login_required
@user_passes_test(is_kitchen_admin)
def get_active_orders_api(request):
    """
    Returns a list of active orders with full details.
    Used for polling by the dashboard to detect and render new orders.
    """
    from django.utils.timesince import timesince
    from django.urls import reverse
    
    orders = Order.objects.filter(
        status__in=[OrderStatus.PENDING, OrderStatus.PREPARING, OrderStatus.READY, OrderStatus.CONFIRMED, OrderStatus.PENDING_FULFILLMENT]
    ).select_related('user__profile').prefetch_related(
        'items__menu_item',
        'items__sweets_item',
        'items__sweets_variation',
        'user__profile__addresses'
    ).distinct().order_by('-created_at')
    
    orders_data = []
    for order in orders:
        # Show all items for this order in the active card
        kitchen_items = order.items.all()
        items_data = []
        for item in kitchen_items[:5]:  # Return up to 5 items for the dashboard
            name = item.item_name_snapshot
            
            if not name:
                if item.menu_item:
                    name = item.menu_item.name
                elif item.sweets_item:
                    name = item.sweets_item.name
                    if item.sweets_variation:
                        name = f"{name} ({item.sweets_variation.weight})"
                elif item.variation_snapshot:
                    name = f"Item ({item.variation_snapshot})"
                else:
                    name = "Item Deleted"
            
            items_data.append({
                'name': name,
                'quantity': item.quantity,
                'week': item.week_number,
                'day': item.day_of_week
            })
        
        # Get customer details
        profile = getattr(order.user, 'profile', None)
        phone = profile.phone_number if profile else ""
        
        address_str = "No address"
        if profile:
            default_address = profile.addresses.filter(is_default=True).first() or profile.addresses.first()
            if default_address:
                address_str = f"{default_address.address_line_1}, {default_address.city}"

        searchable_names = []
        for item in kitchen_items:
            name = item.item_name_snapshot
            if not name:
                if item.menu_item: name = item.menu_item.name
                elif hasattr(item, 'master_item') and item.master_item: name = item.master_item.name
                elif item.sweets_item: name = item.sweets_item.name
            if name: searchable_names.append(name.lower())
            
        orders_data.append({
            'id': order.id,
            'city': order.city,
            'location_name': order.location.name if order.location else None,
            'location_id': order.location_id,
            'searchable_items': "|".join(searchable_names),
            'status': order.status,
            'status_display': order.get_status_display(),
            'created_at': order.created_at.isoformat(),
            'timesince': timesince(order.created_at),
            'pickup_date': str(order.pickup_date) if order.pickup_date else 'Today',
            'pickup_slot': order.pickup_slot.label if order.pickup_slot else None,
            'items_count': kitchen_items.count(),
            'items': items_data,
            'has_meals': kitchen_items.filter(models.Q(menu_item__isnull=False) | models.Q(item_name_snapshot__isnull=False) | models.Q(master_item__isnull=False)).exists(),
            'has_sweets': kitchen_items.filter(sweets_item__isnull=False).exists(), # Sweets are usually NOT wiped
            'detail_url': reverse('kitchen:order_detail', args=[order.id])
        })
    
    return JsonResponse({'orders': orders_data})

# -----------------------------------------------------------
# VENDING PRICES UPDATE
# -----------------------------------------------------------

# @login_required
# @user_passes_test(is_kitchen_admin)
@login_required
@user_passes_test(is_kitchen_admin)
def vending_prices_view(request):
    not_found_items = []
    updated_items = []
    
    if request.method == 'POST' and request.FILES.get('price_file'):
        file = request.FILES['price_file']
        
        try:
            data = []
            if file.name.endswith('.csv'):
                decoded_file = file.read().decode('utf-8-sig').splitlines()
                reader = csv.DictReader(decoded_file)
                data = list(reader)
                
            elif file.name.endswith(('.xls', '.xlsx')):
                import openpyxl
                wb = openpyxl.load_workbook(file)
                sheet = wb.active
                
                rows = list(sheet.iter_rows(values_only=True))
                if rows:
                    headers = [str(h).strip() if h else f"col_{i}" for i, h in enumerate(rows[0])]
                    for row in rows[1:]:
                        row_dict = {}
                        for i, val in enumerate(row):
                            if i < len(headers):
                                row_dict[headers[i]] = val
                        data.append(row_dict)
            else:
                messages.error(request, "Unsupported file format. Please use CSV or Excel.")
                return render(request, 'kitchen/vending_prices.html')

            # Process Data
            if not data:
                with open('debug_log.txt', 'a') as f:
                    f.write("DEBUG: No data found after parsing file.\n")

            # PASS 1: Collect final prices for each item
            price_updates = {} # Map: clean_name -> price_val
            
            for i, row in enumerate(data):
                # Normalize keys slightly to key access
                row_lower = {str(k).lower().strip(): v for k, v in row.items() if k}
                
                if i < 3: # Log first 3 rows
                     with open('debug_log.txt', 'a') as f:
                        f.write(f"DEBUG ROW {i} KEYS: {list(row_lower.keys())}\n")
                        f.write(f"DEBUG ROW {i} VALS: {row_lower}\n")

                # Fetch Name and Price
                # Try 'item', 'item name', 'name'
                raw_name = row_lower.get('item') or row_lower.get('item name') or row_lower.get('name')
                
                # Try 'price', 'new price', 'cost'
                raw_price = row_lower.get('price') or row_lower.get('new price') or row_lower.get('cost')
                
                if not raw_name:
                    continue # Skip empty rows
                
                # Clean Name: Valid name, remove '*', ignore case
                clean_name = str(raw_name).replace('*', '').strip()
                
                # Clean Price
                try:
                     price_val = parse_macros(raw_price) # Reuse parse_macros for float extraction
                except:
                     price_val = 0.0

                if price_val > 0:
                     price_updates[clean_name] = price_val

            # PASS 2: Global Update
            for name, price in price_updates.items():
                qs = MenuItem.objects.filter(name__iexact=name)
                
                if qs.exists():
                    count = qs.update(price=price)
                    updated_items.append({
                        'name': name,
                        'new_price': price,
                        'count': count
                    })
                else:
                    not_found_items.append({
                        'name': name,
                        'price': price
                    })
            
            if updated_items:
                total = sum(item['count'] for item in updated_items)
                messages.success(request, f"Global Update Success: Updated {len(updated_items)} unique items affecting {total} total records.")
            elif not_found_items:
                 messages.warning(request, "Process complete. Some items were not found.")
            else:
                 messages.info(request, "Process complete. No changes needed.")

        except Exception as e:
            messages.error(request, f"Error processing file: {str(e)}")
            
    return render(request, 'kitchen/vending_prices.html', {
        'not_found_items': not_found_items, 
        'updated_items': updated_items
    })

# -----------------------------------------------------------
# VENDING MACHINE ITEMS (STOCK)
# -----------------------------------------------------------

@login_required
@user_passes_test(is_kitchen_admin)
def vending_machine_items_view(request):
    """
    Fetches items from external vending API and structures them by shelf/slot for a visual UI.
    """
    from vending.models import VendingLocation
    
    # 1. Get Active Machines for Selector
    machines = VendingLocation.objects.filter(is_active=True).exclude(serial_number__isnull=True).order_by('name')
    if not machines.exists():
        messages.warning(request, "No active vending machines found in database.")
        return render(request, 'kitchen/vending_machine_items.html', {'machines': [], 'shelves': []})

    # 2. Determine Selected Machine
    selected_uuid = request.GET.get('machine_uuid')
    if not selected_uuid:
        selected_uuid = machines.first().serial_number
    
    current_machine = machines.filter(serial_number=selected_uuid).first()
    
    # 3. Fetch Token
    token_url = "http://www.hnzczy.cn:8087/apiusers/checkusername"
    token_params = {
        "userName": "C202405128888",
        "password": "8888"
    }
    
    shelves_data = []
    
    try:
        token_response = requests.get(token_url, params=token_params, timeout=15)
        token_data = token_response.json()
        token = token_data.get("data") or token_data.get("token")
        
        if not token:
            messages.error(request, "Could not fetch external vending token.")
        else:
            # 4. Fetch Machine Goods
            goods_url = "http://www.hnzczy.cn:8087/commodityinfo/querycommodityinfo"
            stock_url = "http://www.hnzczy.cn:8087/commodityinfo/queryGoodsStock"
            headers = {"Authorization": token}
            
            # Pre-fetch local images to fix broken external URLs
            from vending.models import MenuItem
            import re
            
            def normalize_name(name):
                if not name: return ""
                name = name.replace("&", "and")
                return re.sub(r'[^a-zA-Z0-9]', '', name).lower()

            local_image_map = {}
            for item in MenuItem.objects.exclude(image=''):
                # Map exact name
                if item.image:
                    local_image_map[item.name] = item.image.url
                    # Map normalized name
                    local_image_map[normalize_name(item.name)] = item.image.url
            
            try:
                with concurrent.futures.ThreadPoolExecutor(max_workers=2) as executor:
                    future_goods = executor.submit(requests.get, goods_url, params={"machineUuid": selected_uuid}, headers=headers, timeout=25)
                    future_stock = executor.submit(requests.get, stock_url, params={"machineUuid": selected_uuid}, headers=headers, timeout=25)

                    goods_res = future_goods.result()
                    stock_res = future_stock.result()
                
                api_data = goods_res.json() # This was 'response.json()' before
                stock_data = stock_res.json() # New stock data

                # Process goods data
                if api_data and api_data.get("result") == "200":
                    slots = api_data.get("data") or []
                    shelves_dict = {}
                    
                    for slot in slots:
                        # Group by Tier (Shelf)
                        raw_tier = slot.get("modityTierSeq", 0)
                        try:
                            tier = int(raw_tier)
                        except (ValueError, TypeError):
                            tier = 999 

                        if tier not in shelves_dict:
                            shelves_dict[tier] = {
                                'id': tier,
                                'name': f"Shelf {tier}",
                                'spots_map': {} # Use dict for strict deduplication
                            }
                        
                        goods = slot.get("commGoodsResp")
                        
                        # Normalize Slot ID
                        raw_spot_id = slot.get('arrivalName')
                        spot_id_str = str(raw_spot_id).strip()
                        
                        spot_data = {
                            'arrivalName': spot_id_str,
                            'modityTierNum': slot.get('modityTierNum'), 
                            'capacity': slot.get('arrivalCapacity'),
                            'present': slot.get('presentNumber', 0),
                            'status': 'empty',
                            'item': None
                        }
                        
                        if goods:
                            g_name = goods.get('goodsName')
                            local_img = local_image_map.get(g_name) or local_image_map.get(normalize_name(g_name))
                            
                            spot_data['item'] = {
                                'uuid': goods.get('uuid'),
                                'name': g_name,
                                'price': goods.get('goodsPrice'),
                                'image': local_img if local_img else goods.get('goodsUrl'),
                                'desc': goods.get('goodsDesc')
                            }
                            
                            if slot.get('presentNumber', 0) > 0:
                                spot_data['status'] = 'available'
                            else:
                                spot_data['status'] = 'sold_out'

                        # Deduplication / Merge Logic
                        # If spot exists, only overwrite if current one has item (is better)
                        existing = shelves_dict[tier]['spots_map'].get(spot_id_str)
                        if existing:
                            if spot_data['item'] is not None:
                                shelves_dict[tier]['spots_map'][spot_id_str] = spot_data
                            # Else: keep existing (which might have item or be empty, doesn't matter, we prioritize occupied)
                        else:
                             shelves_dict[tier]['spots_map'][spot_id_str] = spot_data
                    
                    # Convert maps to sorted lists
                    sorted_tiers = sorted(shelves_dict.keys())
                    for tier in sorted_tiers:
                        shelf = shelves_dict[tier]
                        # Flatten spots map values
                        spots_list = list(shelf['spots_map'].values())
                        
                        # Sort spots
                        spots_list.sort(key=lambda x: (int(x['modityTierNum']) if str(x['modityTierNum']).isdigit() else 999, x['arrivalName']))
                        
                        shelf['spots'] = spots_list
                        shelves_data.append(shelf)
                        
                else:
                    messages.warning(request, f"API returned error or empty data: {api_data.get('msg', 'Unknown Error')}")

            except Exception as e:
                print(f"Error fetching machine data: {e}")
                messages.error(request, f"Failed to connect to machine API: {str(e)}")

    except Exception as e:
        messages.error(request, f"Error fetching token: {str(e)}")

    context = {
        'machines': machines,
        'current_machine_uuid': selected_uuid,
        'current_machine': current_machine,
        'shelves': shelves_data
    }
    
    return render(request, 'kitchen/vending_machine_items.html', context)

@login_required
@user_passes_test(is_kitchen_admin)
@require_POST
def update_vending_stock(request):
    """
    Updates the quantity of a vending item.
    """
    uuid = request.POST.get('uuid')
    quantity = request.POST.get('quantity')
    
    if uuid and quantity is not None:
        try:
            stock = VendingMachineStock.objects.get(vending_good_uuid=uuid)
            stock.quantity = int(quantity)
            stock.save()
            messages.success(request, f"Updated stock for {stock.goods_name}")
        except Exception as e:
            messages.error(request, f"Error updating stock: {str(e)}")
            
    machine_uuid = request.POST.get('machine_uuid')
    
    response = redirect('kitchen:vending_machine_items')
    if machine_uuid:
        response['Location'] += f'?machine_uuid={machine_uuid}'
    return response

# @login_required
# @user_passes_test(is_kitchen_admin)
@login_required
@user_passes_test(is_kitchen_admin)
def daily_orders_view(request):
    """
    Shows aggregated orders for a specific day.
    Includes 'Order Now' items for that date and 'Weekly/Monthly' plan items for that day of week.
    """
    date_str = request.GET.get('date')
    if date_str:
        try:
            selected_date = timezone.datetime.strptime(date_str, '%Y-%m-%d').date()
        except ValueError:
            selected_date = timezone.now().date()
    else:
        selected_date = timezone.now().date()
    
    day_name = selected_date.strftime('%A') # e.g. "Tuesday"
    
    # Query items for this day
    # Condition 1: Immediate orders (Order Now / Smart Grab) scheduled for this date
    # Condition 2: Plan orders for this day of week
    
    # We need to capture:
    # A. Plan items (Weekly/Monthly) matching day_of_week
    # B. Daily items (Order Now/Smart Grab) matching pickup_date (or created_today if pickup_date is used)
    
    # Fetch Items (Sorted by Time, then Name)
    items = OrderItem.objects.filter(
        (
            Q(plan_type='START_PLAN', plan_subtype__in=[PlanSubType.WEEKLY, PlanSubType.MONTHLY], day_of_week=day_name)
        ) | (
            Q(plan_type__in=['ORDER_NOW', 'SMART_GRAB'], pickup_date=selected_date)
        )
    ).filter(
        pickup_code__isnull=True, 
        order__status__in=[
            OrderStatus.CONFIRMED, 
            OrderStatus.PREPARING, 
            OrderStatus.READY,
            OrderStatus.PENDING 
        ]
    ).select_related('menu_item', 'order', 'order__user', 'pickup_slot', 'order__pickup_slot').order_by('order__user__username', 'week_number', 'pickup_slot__start_time', 'menu_item__name')

    # Separate into Pending and Ready
    pending_items = []
    ready_items = []
    
    for item in items:
        # Determine slot label for display
        slot_obj = item.pickup_slot or item.order.pickup_slot
        item.slot_label = slot_obj.label if slot_obj else "Standard Pickup"
        item.user_name = f"{item.order.user.first_name} {item.order.user.last_name}".strip() or item.order.user.username
        
        if item.status == OrderStatus.READY:
            ready_items.append(item)
        else:
            pending_items.append(item)

    context = {
        'selected_date': selected_date,
        'day_name': day_name,
        'pending_items': pending_items,
        'ready_items': ready_items,
        'date_str': selected_date.strftime('%Y-%m-%d'),
        'total_orders_count': items.count(),
        'pending_count': len(pending_items),
        'ready_count': len(ready_items)
    }
    return render(request, 'kitchen/daily_orders.html', context)

# ==========================================
# MASTER ITEMS MANAGEMENT (VENDING & CATERING)
# ==========================================
from vending.models import MasterItem as VendingMasterItem
from catering.models import CateringMasterItem
from django.views.generic import ListView

class VendingMasterListView(LoginRequiredMixin, ListView):
    model = VendingMasterItem
    template_name = 'kitchen/master_items_vending.html'
    context_object_name = 'items'
    ordering = ['name']

    def get_queryset(self):
        from django.db.models import Count, Q
        from vending.models import MenuType
        queryset = super().get_queryset()
        query = self.request.GET.get('q')
        if query:
            queryset = queryset.filter(name__icontains=query)
            
        queryset = queryset.annotate(
            standard_schedules_count=Count(
                'menu_items', 
                filter=Q(menu_items__menu__menu_type=MenuType.STANDARD),
                distinct=True
            ),
            weekly_schedules_count=Count(
                'menu_items',
                filter=Q(menu_items__menu__menu_type=MenuType.WEEKLY),
                distinct=True
            ),
            monthly_schedules_count=Count(
                'menu_items',
                filter=Q(menu_items__menu__menu_type=MenuType.MONTHLY),
                distinct=True
            )
        )
        return queryset

    def post(self, request, *args, **kwargs):
        # Quick inline update handler
        item_id = request.POST.get('item_id')
        new_name = request.POST.get('name')
        new_desc = request.POST.get('description')
        
        if item_id and new_name:
            try:
                item = VendingMasterItem.objects.get(id=item_id)
                item.name = new_name
                if new_desc is not None:
                    item.description = new_desc
                if 'image' in request.FILES:
                    item.image = request.FILES['image']
                
                item.save() # This triggers the signal to update all MenuItems!
                messages.success(request, f"Updated '{item.name}' successfully.")
            except Exception as e:
                messages.error(request, f"Error updating item: {e}")
        
        return redirect('kitchen:vending_master_list')

class CateringMasterListView(LoginRequiredMixin, ListView):
    model = CateringMasterItem
    template_name = 'kitchen/master_items_catering.html'
    context_object_name = 'items'
    ordering = ['name']

    def get_queryset(self):
        queryset = super().get_queryset()
        query = self.request.GET.get('q')
        if query:
            queryset = queryset.filter(name__icontains=query)
        return queryset

    def post(self, request, *args, **kwargs):
        form_mode = request.POST.get('form_mode', 'edit')
        item_id = request.POST.get('item_id')

        # ── DELETE ───────────────────────────────────────────────────────────
        if form_mode == 'delete':
            if item_id:
                try:
                    item = CateringMasterItem.objects.get(id=item_id)
                    name = item.name
                    item.delete()
                    messages.success(request, f"'{name}' deleted successfully.")
                except CateringMasterItem.DoesNotExist:
                    messages.error(request, "Item not found.")
                except Exception as e:
                    messages.error(request, f"Could not delete item: {e}")
            return redirect('kitchen:catering_master_list')

        # ── EDIT ─────────────────────────────────────────────────────────────
        new_name = request.POST.get('name', '').strip()
        new_desc = request.POST.get('description', '')

        if item_id and new_name:
            try:
                item = CateringMasterItem.objects.get(id=item_id)
                item.name = new_name
                item.description = new_desc
                if 'image' in request.FILES:
                    item.image = request.FILES['image']
                item.save()
                messages.success(request, f"Updated '{item.name}' successfully.")
            except CateringMasterItem.DoesNotExist:
                messages.error(request, "Item not found.")
            except Exception as e:
                if 'unique' in str(e).lower() or 'UNIQUE' in str(e):
                    messages.error(request, f"A catering item named '{new_name}' already exists.")
                else:
                    messages.error(request, f"Error updating item: {e}")

        return redirect('kitchen:catering_master_list')


# -----------------------------------------------------------
# CATERING MASTER ITEM — CREATE
# -----------------------------------------------------------

@login_required
@user_passes_test(is_kitchen_admin)
@require_POST
def catering_master_item_create_view(request):
    """Creates a new CateringMasterItem from the Add Item modal."""
    from catering.models import CateringMasterItem
    name = request.POST.get('name', '').strip()
    if not name:
        messages.error(request, "Item name is required.")
        return redirect('kitchen:catering_master_list')

    item = CateringMasterItem(
        name=name,
        description=request.POST.get('description', ''),
    )
    if 'image' in request.FILES:
        item.image = request.FILES['image']

    try:
        item.save()
        messages.success(request, f"Catering item '{item.name}' created successfully.")
    except Exception as e:
        if 'unique' in str(e).lower() or 'UNIQUE' in str(e):
            messages.error(request, f"A catering item named '{name}' already exists. Please choose a different name.")
        else:
            messages.error(request, f"Could not create item: {e}")
    return redirect('kitchen:catering_master_list')


# -----------------------------------------------------------
# SYNC MASTER ITEMS (Maintenance Tools)
# -----------------------------------------------------------

@login_required
@user_passes_test(is_kitchen_admin)
def sync_vending_master_items(request):
    """
    Iterates through all Vending MenuItems and saves them.
    This triggers the 'pre_save' signal to link/create MasterItems.
    """
    from vending.models import MenuItem
    
    count = 0
    items = MenuItem.objects.all()
    for item in items:
        # Just saving triggers the signal
        item.save()
        count += 1
        
    messages.success(request, f"Successfully synced {count} Vending items to Master list.")
    return redirect('kitchen:vending_master_list')


@login_required
@user_passes_test(is_kitchen_admin)
def sync_catering_master_items(request):
    """
    Iterates through all Catering Items (all types) and saves them.
    This triggers the 'pre_save' signal to link/create CateringMasterItems.
    """
    from catering.models import (
        MenuItem as CateringMenuItem,
        CoffeeBreakItem,
        PlatterItem,
        BoxedMealItem,
        LiveStationItem,
        AmericanMenuItem,
        CanapeItem
    )
    
    total_count = 0
    
    # 1. Standard Menu Items
    items = CateringMenuItem.objects.all()
    for item in items:
        item.save()
        total_count += 1
        
    # 2. Coffee Break
    items = CoffeeBreakItem.objects.all()
    for item in items:
        item.save()
        total_count += 1

    # 3. Platters
    items = PlatterItem.objects.all()
    for item in items:
        item.save()
        total_count += 1

    # 4. Boxed Meals
    items = BoxedMealItem.objects.all()
    for item in items:
        item.save()
        total_count += 1

    # 5. Live Stations
    items = LiveStationItem.objects.all()
    for item in items:
        item.save()
        total_count += 1

    # 6. American Menu
    items = AmericanMenuItem.objects.all()
    for item in items:
        item.save()
        total_count += 1

    # 7. Canapes
    items = CanapeItem.objects.all()
    for item in items:
        item.save()
        total_count += 1
        
    messages.success(request, f"Successfully synced {total_count} Catering items to Master list.")
    return redirect('kitchen:catering_master_list')


# -----------------------------------------------------------
# AUTH / LOGOUT
# -----------------------------------------------------------
from django.contrib.auth import logout

def kitchen_logout_view(request):
    """
    Logs out the user and redirects to the admin login page.
    """
    logout(request)
    messages.info(request, "You have been logged out.")
    return redirect('/admin/login/')


@login_required
@user_passes_test(is_kitchen_admin)
def kitchen_retry_fulfillment(request, order_id):
    """
    POST /kitchen/order/<order_id>/retry-fulfillment/
    Kitchen admin retries pickup code generation for a PENDING_FULFILLMENT order.
    """
    if request.method != 'POST':
        return redirect('kitchen:tracking')

    try:
        order = Order.objects.get(id=order_id)
    except Order.DoesNotExist:
        messages.error(request, f"Order #{order_id} not found.")
        return redirect('kitchen:tracking')

    if order.pickup_code:
        messages.info(request, f"Order #{order_id} already has pickup code: {order.pickup_code}")
        return redirect('kitchen:tracking')

    MAX_ATTEMPTS = 10  # Higher limit for admin
    if order.fulfillment_attempts >= MAX_ATTEMPTS:
        messages.error(request, f"Order #{order_id} has exceeded maximum retry attempts ({MAX_ATTEMPTS}).")
        return redirect('kitchen:tracking')

    from vending.services import VendingService
    order.fulfillment_attempts += 1
    order.save(update_fields=['fulfillment_attempts'])

    pickup_code = None
    try:
        pickup_code = VendingService.process_order_fulfillment(order)
    except Exception as e:
        messages.error(request, f"Fulfillment error for Order #{order_id}: {str(e)}")
        return redirect('kitchen:tracking')

    if pickup_code:
        order.pickup_code = pickup_code
        order.qr_code_url = f"https://api.qrserver.com/v1/create-qr-code/?size=150x150&data={pickup_code}"
        order.status = OrderStatus.READY
        order.save(update_fields=['pickup_code', 'qr_code_url', 'status'])
        from vending.models import PlanType
        order.items.filter(plan_type__in=[PlanType.ORDER_NOW, PlanType.SMART_GRAB]).update(
            status=OrderStatus.READY, pickup_code=pickup_code
        )
        messages.success(request, f"✅ Order #{order_id} fulfilled. Pickup code: {pickup_code}")
    else:
        order.status = OrderStatus.PENDING_FULFILLMENT
        order.save(update_fields=['status'])
        messages.error(request, f"❌ Fulfillment failed again for Order #{order_id}. Attempt #{order.fulfillment_attempts}.")

    return redirect('kitchen:tracking')


@login_required
@user_passes_test(is_kitchen_admin)
def kitchen_mark_qr_used(request, order_id):
    """
    POST /kitchen/order/<order_id>/mark-qr-used/
    Kitchen admin marks QR as used (food dispensed), completing the order.
    """
    if request.method != 'POST':
        return redirect('kitchen:tracking')

    try:
        order = Order.objects.get(id=order_id)
    except Order.DoesNotExist:
        messages.error(request, f"Order #{order_id} not found.")
        return redirect('kitchen:tracking')

    if order.qr_used:
        messages.info(request, f"Order #{order_id} QR already marked as used.")
        return redirect('kitchen:tracking')

    order.qr_used = True
    order.status = OrderStatus.COMPLETED
    order.save(update_fields=['qr_used', 'status'])
    from vending.models import PlanType
    order.items.filter(plan_type__in=[PlanType.ORDER_NOW, PlanType.SMART_GRAB]).update(
        status=OrderStatus.COMPLETED
    )
    messages.success(request, f"✅ Order #{order_id} marked as delivered.")
    return redirect('kitchen:tracking')

@login_required
@user_passes_test(is_kitchen_admin)
def agent_dashboard_api(request):
    """
    JSON API for the Agent Dashboard.
    """
    from ai_agents.models import CustomerInquiry, CustomerServiceTicket, AgentActivity, AgentInteractionLog, WhatsAppDevice
    
    # Fetch real inquiries from DB (Sales Agent)
    recent_inquiries = CustomerInquiry.objects.all().order_by('-timestamp')[:50]
    inquiries = [{
        "Name": inquiry.name,
        "Email": inquiry.email,
        "Phone": inquiry.phone,
        "Preference": inquiry.preference,
        "Date": inquiry.event_date,
        "Time": inquiry.event_time,
        "Number of People": inquiry.people_count,
        "Venue": inquiry.venue,
        "timestamp": inquiry.timestamp.strftime("%Y-%m-%d %H:%M")
    } for inquiry in recent_inquiries]

    # Fetch real tickets from DB (CS Agent)
    recent_tickets = CustomerServiceTicket.objects.all().order_by('-timestamp')[:30]
    tickets = [{
        "phone": ticket.phone,
        "subject": ticket.subject,
        "message": ticket.message,
        "status": ticket.status,
        "timestamp": ticket.timestamp.strftime("%Y-%m-%d %H:%M")
    } for ticket in recent_tickets]

    # Define the base layout of agents
    agent_defs = [
        {"id": "sales", "name": "Sales Agent", "color": "blue"},
        {"id": "customer_service", "name": "Customer Service", "color": "green"},
        {"id": "lead_gen", "name": "Lead Generation", "color": "amber"},
        {"id": "marketing", "name": "Marketing Agent", "color": "purple"},
        {"id": "accounting", "name": "Accounting Assistant", "color": "slate"},
    ]
    
    agents = []
    for adef in agent_defs:
        # Fetch real activity or use defaults if not yet active
        activity = AgentActivity.objects.filter(agent_id=adef['id']).first()
        agents.append({
            "id": adef['id'],
            "name": adef['name'],
            "status": activity.status if activity else "Active",
            "is_paused": activity.is_paused if activity else False,
            "last_activity": activity.last_task if activity and activity.last_task else "Standing by...",
            "color": adef['color'],
            "updated_at": activity.updated_at.strftime("%H:%M") if activity else None
        })
        
    # Fetch recent transcripts
    recent_logs = AgentInteractionLog.objects.all()[:20]
    transcripts = [{
        "agent_id": log.agent_id,
        "user_message": log.user_message,
        "ai_response": log.ai_response,
        "timestamp": log.timestamp.strftime("%Y-%m-%d %H:%M:%S")
    } for log in recent_logs]

    # Fetch recent scraped leads
    from ai_agents.models import ScrapedLead
    from django.db.models import Count
    
    # Calculate global stats for all leads in the database
    lead_stats_query = ScrapedLead.objects.values('status').annotate(total=Count('status'))
    lead_stats = {
        'New': 0,
        'Contacted': 0,
        'Failed': 0
    }
    for item in lead_stats_query:
        lead_stats[item['status']] = item['total']

    recent_leads = ScrapedLead.objects.all().order_by('-date_scraped')[:150]
    leads = [{
        "company_name": lead.company_name,
        "phone_number": lead.phone_number,
        "status": lead.status,
        "source": lead.source,
        "date_scraped": lead.date_scraped.strftime("%Y-%m-%d %H:%M")
    } for lead in recent_leads]

    # Fetch WhatsApp Gateway status
    device, _ = WhatsAppDevice.objects.get_or_create(id=1)
    whatsapp_gateway = {
        "status": device.status,
        "phone_number": device.phone_number,
        "qr_code": device.qr_code,
        "updated_at": device.updated_at.strftime("%Y-%m-%d %H:%M") if device.updated_at else None
    }

    # Fetch recent Ad Drafts
    from ai_agents.models import AdDraft
    recent_ad_drafts = AdDraft.objects.all().order_by('-created_at')[:10]
    ad_drafts = [{
        "id": draft.id,
        "platform": draft.platform,
        "headline": draft.headline,
        "body": draft.body_text,
        "targeting": draft.targeting_summary,
        "budget": str(draft.budget),
        "status": draft.status,
        "platform_ad_id": draft.platform_ad_id,
        "rejection_reason": draft.rejection_reason,
        "created_at": draft.created_at.strftime("%b %d, %H:%M")
    } for draft in recent_ad_drafts]

    # Fetch latest marketing report
    from ai_agents.models import MarketingReport
    latest_report = MarketingReport.objects.order_by('-date', '-created_at').first()
    marketing_data = None
    if latest_report:
        marketing_data = {
            "date": latest_report.date.strftime("%Y-%m-%d"),
            "meta_spend": f"{latest_report.meta_spend:,.2f} AED",
            "google_spend": f"{latest_report.google_spend:,.2f} AED",
            "ai_analysis": latest_report.ai_analysis_text
        }

    return JsonResponse({
        'inquiries': inquiries,
        'tickets': tickets,
        'agents': agents,
        'transcripts': transcripts,
        'leads': leads,
        'lead_stats': lead_stats,
        'whatsapp_gateway': whatsapp_gateway,
        'maton_configured': bool(os.getenv('MATON_API_KEY')),
        'marketing_report': marketing_data,
        'ad_drafts': ad_drafts,
    })

@login_required
@user_passes_test(is_kitchen_admin)
@require_POST
def approve_ad_draft(request, pk):
    """Approves an ad draft and triggers publication to the platform."""
    from ai_agents.models import AdDraft
    from ai_agents.services.meta_ads_service import MetaAdsService
    # (Other platform services would be imported here)
    
    draft = get_object_or_404(AdDraft, pk=pk)
    
    if draft.status != 'Pending Approval':
        messages.error(request, "This ad is not in a pending state.")
        return redirect('kitchen:agent_dashboard')
    
    draft.status = 'Approved'
    draft.save()
    
    # Trigger API Call
    success = False
    platform_id = None
    
    if draft.platform == 'Meta':
        success, platform_id = MetaAdsService.publish_ad(draft)
    else:
        # Placeholder for Google
        success, platform_id = False, "Google Ads automatic publishing not yet implemented."

    if success:
        draft.status = 'Live'
        draft.platform_ad_id = platform_id
        messages.success(request, f"Ad published successfully to {draft.platform}! ID: {platform_id}")
    else:
        draft.status = 'Failed'
        messages.error(request, f"Failed to publish ad: {platform_id}")
    
    draft.save()
    return redirect('kitchen:agent_dashboard')

@login_required
@user_passes_test(is_kitchen_admin)
@require_POST
def reject_ad_draft(request, pk):
    """Rejects an ad draft with an optional reason."""
    from ai_agents.models import AdDraft
    draft = get_object_or_404(AdDraft, pk=pk)
    reason = request.POST.get('reason', 'Rejected by Admin')
    
    draft.status = 'Rejected'
    draft.rejection_reason = reason
    draft.save()
    
    messages.warning(request, f"Ad draft for '{draft.headline}' was rejected.")
    return redirect('kitchen:agent_dashboard')

@login_required
@user_passes_test(is_kitchen_admin)
def agent_dashboard_view(request):
    """
    Shell view for the Agent Dashboard.
    """
    return render(request, 'kitchen/agent_dashboard.html')

# ==========================================
# WEEKLY AND MONTHLY ASSIGNMENT VIEWS
# ==========================================
from django.db import transaction
from vending.models import MenuType

class WeeklyVendingAssignmentView(LoginRequiredMixin, UserPassesTestMixin, TemplateView):
    template_name = 'kitchen/weekly_vending_assignment.html'

    def test_func(self):
        return is_kitchen_admin(self.request.user)

    def get_context_data(self, **kwargs):
        from vending.models import MenuItem
        context = super().get_context_data(**kwargs)
        days = ['Monday', 'Tuesday', 'Wednesday', 'Thursday', 'Friday']
        selected_day = self.request.GET.get('day', 'Monday')
        if selected_day not in days:
            selected_day = 'Monday'
            
        q = self.request.GET.get('q', '')
        
        items = VendingMasterItem.objects.all().order_by('name')
        if q:
            items = items.filter(name__icontains=q)
            
        # Get active assignments for this day on Weekly menu
        active_master_ids = set(MenuItem.objects.filter(
            menu__day_of_week=selected_day,
            menu__menu_type=MenuType.WEEKLY
        ).values_list('master_item_id', flat=True))
        
        for item in items:
            item.is_assigned = item.id in active_master_ids

        # Sort so assigned items come first
        items = sorted(items, key=lambda x: (not x.is_assigned, x.name))
            
        context['days'] = days
        context['selected_day'] = selected_day
        context['items'] = items
        return context

@login_required
@user_passes_test(is_kitchen_admin)
@require_POST
def save_weekly_vending_assignment(request):
    from vending.models import Menu, MenuItem
    valid_days = ['Monday', 'Tuesday', 'Wednesday', 'Thursday', 'Friday']

    # Support multi-day submission (days[] + item_ids_{day} per day)
    days_submitted = request.POST.getlist('days')
    if not days_submitted:
        # Legacy single-day fallback
        days_submitted = [request.POST.get('day', 'Monday')]

    active_day = request.POST.get('active_day', days_submitted[0] if days_submitted else 'Monday')

    with transaction.atomic():
        for day in days_submitted:
            if day not in valid_days:
                continue
            item_ids = request.POST.getlist(f'item_ids_{day}')

            menu, _ = Menu.objects.get_or_create(
                day_of_week=day,
                menu_type=MenuType.WEEKLY,
                defaults={'week_number': 1}
            )
            menu.items.all().delete()

            masters = VendingMasterItem.objects.filter(id__in=item_ids)
            for master in masters:
                MenuItem.objects.create(
                    menu=menu,
                    master_item=master,
                    name=master.name,
                    price=master.default_price,
                    description=master.description,
                    image=master.image,
                    image_source_url=master.image_source_url
                )

    days_label = ', '.join(days_submitted)
    messages.success(request, f"Successfully updated Weekly menu for: {days_label}.")
    return redirect(f"{reverse('kitchen:vending_weekly_assignment')}?day={active_day}")

class MonthlyVendingAssignmentView(LoginRequiredMixin, UserPassesTestMixin, TemplateView):
    template_name = 'kitchen/monthly_vending_assignment.html'

    def test_func(self):
        return is_kitchen_admin(self.request.user)

    def get_context_data(self, **kwargs):
        from vending.models import MenuItem
        context = super().get_context_data(**kwargs)
        days = ['Monday', 'Tuesday', 'Wednesday', 'Thursday', 'Friday']
        weeks = [1, 2, 3, 4]
        
        selected_day = self.request.GET.get('day', 'Monday')
        if selected_day not in days:
            selected_day = 'Monday'
            
        try:
            selected_week = int(self.request.GET.get('week', 1))
            if selected_week not in weeks:
                selected_week = 1
        except ValueError:
            selected_week = 1
            
        q = self.request.GET.get('q', '')
        
        items = VendingMasterItem.objects.all().order_by('name')
        if q:
            items = items.filter(name__icontains=q)
            
        # Get active assignments for this day and week on Monthly menu
        active_master_ids = set(MenuItem.objects.filter(
            menu__day_of_week=selected_day,
            menu__week_number=selected_week,
            menu__menu_type=MenuType.MONTHLY
        ).values_list('master_item_id', flat=True))
        
        for item in items:
            item.is_assigned = item.id in active_master_ids

        # Sort so assigned items come first
        items = sorted(items, key=lambda x: (not x.is_assigned, x.name))
            
        context['days'] = days
        context['weeks'] = weeks
        context['selected_day'] = selected_day
        context['selected_week'] = selected_week
        context['items'] = items
        return context

@login_required
@user_passes_test(is_kitchen_admin)
@require_POST
def save_monthly_vending_assignment(request):
    from vending.models import Menu, MenuItem
    valid_days = ['Monday', 'Tuesday', 'Wednesday', 'Thursday', 'Friday']
    valid_weeks = [1, 2, 3, 4]

    # Support multi-tab submission: tabs[] and item_ids_{week}_{day} per combination
    tabs_submitted = request.POST.getlist('tabs')
    active_day = request.POST.get('active_day', 'Monday')
    try:
        active_week = int(request.POST.get('active_week', 1))
    except ValueError:
        active_week = 1

    if not tabs_submitted:
        # Legacy single-tab fallback
        day = request.POST.get('day', 'Monday')
        try:
            week = int(request.POST.get('week', 1))
        except ValueError:
            week = 1
        tabs_submitted = [f'{week}_{day}']

    with transaction.atomic():
        for tab_key in tabs_submitted:
            try:
                week_str, day = tab_key.split('_', 1)
                week = int(week_str)
            except (ValueError, AttributeError):
                continue
            if day not in valid_days or week not in valid_weeks:
                continue

            item_ids = request.POST.getlist(f'item_ids_{tab_key}')

            menu, _ = Menu.objects.get_or_create(
                day_of_week=day,
                week_number=week,
                menu_type=MenuType.MONTHLY
            )
            menu.items.all().delete()

            masters = VendingMasterItem.objects.filter(id__in=item_ids)
            for master in masters:
                MenuItem.objects.create(
                    menu=menu,
                    master_item=master,
                    name=master.name,
                    price=master.default_price,
                    description=master.description,
                    image=master.image,
                    image_source_url=master.image_source_url
                )

    messages.success(request, f"Successfully saved Monthly menu assignments.")
    return redirect(f"{reverse('kitchen:vending_monthly_assignment')}?day={active_day}&week={active_week}")


# -----------------------------------------------------------
# VENDING LOCATIONS MANAGEMENT
# -----------------------------------------------------------

@login_required
@user_passes_test(is_kitchen_admin)
def locations_view(request):
    """Lists all VendingLocations."""
    from vending.models import VendingLocation
    locations = VendingLocation.objects.all().order_by('name')
    return render(request, 'kitchen/locations.html', {
        'locations': locations,
        'total_count': locations.count(),
        'active_count': locations.filter(is_active=True).count(),
        'inactive_count': locations.filter(is_active=False).count(),
    })


@login_required
@user_passes_test(is_kitchen_admin)
def locations_manage(request):
    """Handles add / edit / delete of a VendingLocation via POST."""
    from vending.models import VendingLocation

    if request.method != 'POST':
        return redirect('kitchen:locations')

    action = request.POST.get('action')

    # ── DELETE ────────────────────────────────────────────────────────
    if action == 'delete':
        loc = get_object_or_404(VendingLocation, pk=request.POST.get('location_id'))
        name = loc.name
        loc.delete()
        messages.success(request, f"Location '{name}' deleted successfully.")
        return redirect('kitchen:locations')

    # ── ADD / EDIT ────────────────────────────────────────────────────
    name = request.POST.get('name', '').strip()
    info = request.POST.get('info', '').strip()
    hours = request.POST.get('hours', '').strip()
    serial = request.POST.get('serial_number', '').strip() or None
    is_active = request.POST.get('is_active') == 'true'

    try:
        lat = float(request.POST.get('latitude', 0))
        lng = float(request.POST.get('longitude', 0))
    except (ValueError, TypeError):
        messages.error(request, "Invalid latitude or longitude value.")
        return redirect('kitchen:locations')

    if not name:
        messages.error(request, "Location name is required.")
        return redirect('kitchen:locations')

    if action == 'add':
        VendingLocation.objects.create(
            name=name, info=info, hours=hours,
            latitude=lat, longitude=lng,
            serial_number=serial, is_active=is_active
        )
        messages.success(request, f"Location '{name}' added successfully.")

    elif action == 'edit':
        loc = get_object_or_404(VendingLocation, pk=request.POST.get('location_id'))
        loc.name = name
        loc.info = info
        loc.hours = hours
        loc.latitude = lat
        loc.longitude = lng
        loc.serial_number = serial
        loc.is_active = is_active
        loc.save()
        messages.success(request, f"Location '{name}' updated successfully.")

    return redirect('kitchen:locations')


# -----------------------------------------------------------
# LOCATION-BASED PRICES (Per-Machine Pricing)
# -----------------------------------------------------------

@login_required
@user_passes_test(is_kitchen_admin)
def location_based_prices_view(request):
    """
    Admin tab where prices can be overridden per machine/location.
    - No location selected → renders machine picker.
    - Location selected (?location_id=) → renders all vending master items
      with their effective price for that location and any override badge.
    """
    from vending.models import VendingLocation, MasterItem, LocationItemPrice

    locations = VendingLocation.objects.filter(is_active=True).order_by('name')

    raw_loc = request.GET.get('location_id')
    try:
        location_id = int(raw_loc) if raw_loc else None
    except (TypeError, ValueError):
        location_id = None

    selected_location = None
    items_data = []
    overrides_count = 0

    if location_id:
        selected_location = locations.filter(id=location_id).first()
        if not selected_location:
            messages.error(request, "Selected machine not found.")
            return redirect('kitchen:location_based_prices')

        query = (request.GET.get('q') or '').strip()

        overrides_qs = LocationItemPrice.objects.filter(location=selected_location)
        overrides_map = {o.master_item_id: o for o in overrides_qs}
        overrides_count = len(overrides_map)

        items_qs = MasterItem.objects.all()
        if query:
            items_qs = items_qs.filter(name__icontains=query)

        for item in items_qs:
            override = overrides_map.get(item.id)
            items_data.append({
                'id': item.id,
                'name': item.name,
                'default_price': item.default_price,
                'effective_price': override.price if override else item.default_price,
                'has_override': override is not None,
                'override_updated_at': override.updated_at if override else None,
                'image_url': item.image.url if item.image else None,
            })

        # SORTING: Items with recent overrides first, then by name
        # We use a large dummy date for items without overrides to sort them after
        from datetime import datetime
        items_data.sort(key=lambda x: (
            0 if x['has_override'] else 1,
            -x['override_updated_at'].timestamp() if x['override_updated_at'] else 0,
            x['name']
        ))

    return render(request, 'kitchen/location_based_prices.html', {
        'locations': locations,
        'selected_location': selected_location,
        'items': items_data,
        'overrides_count': overrides_count,
        'search_query': request.GET.get('q', ''),
    })


@login_required
@user_passes_test(is_kitchen_admin)
@require_POST
def location_price_set(request):
    """
    POST: location_id, master_item_id, price → creates or updates an override.
    """
    from vending.models import VendingLocation, MasterItem, LocationItemPrice
    from decimal import Decimal, InvalidOperation

    location_id = request.POST.get('location_id')
    master_item_id = request.POST.get('master_item_id')
    raw_price = (request.POST.get('price') or '').strip()

    if not (location_id and master_item_id and raw_price):
        messages.error(request, "Missing location, item, or price.")
        return redirect(f"{reverse('kitchen:location_based_prices')}?location_id={location_id or ''}")

    try:
        price = Decimal(raw_price)
    except (InvalidOperation, ValueError):
        messages.error(request, f"Invalid price: {raw_price}")
        return redirect(f"{reverse('kitchen:location_based_prices')}?location_id={location_id}")

    if price < 0:
        messages.error(request, "Price cannot be negative.")
        return redirect(f"{reverse('kitchen:location_based_prices')}?location_id={location_id}")

    location = get_object_or_404(VendingLocation, pk=location_id)
    master = get_object_or_404(MasterItem, pk=master_item_id)

    obj, created = LocationItemPrice.objects.update_or_create(
        location=location,
        master_item=master,
        defaults={'price': price},
    )

    msg = f"Set price for '{master.name}' at {location.name}: AED {price}" if created else f"Updated price for '{master.name}' at {location.name}: AED {price}"
    
    if request.headers.get('X-Requested-With') == 'XMLHttpRequest':
        return JsonResponse({
            'status': 'success',
            'message': msg,
            'effective_price': str(price),
            'has_override': True,
            'updated_at': obj.updated_at.strftime("%b %d, %H:%M")
        })

    messages.success(request, msg)
    return redirect(f"{reverse('kitchen:location_based_prices')}?location_id={location_id}")


@login_required
@user_passes_test(is_kitchen_admin)
@require_POST
def location_price_clear(request):
    """
    POST: location_id, master_item_id → removes the override (reverts to master default).
    """
    from vending.models import LocationItemPrice

    location_id = request.POST.get('location_id')
    master_item_id = request.POST.get('master_item_id')

    if not (location_id and master_item_id):
        messages.error(request, "Missing location or item.")
        return redirect('kitchen:location_based_prices')

    LocationItemPrice.objects.filter(
        location_id=location_id, master_item_id=master_item_id
    ).delete()

    msg = "Price override removed. Master default will be used."
    if request.headers.get('X-Requested-With') == 'XMLHttpRequest':
        # Need to find the master item to return its default price
        from vending.models import MasterItem
        master = MasterItem.objects.get(pk=master_item_id)
        return JsonResponse({
            'status': 'success',
            'message': msg,
            'effective_price': str(master.default_price),
            'has_override': False
        })

    messages.success(request, msg)
    return redirect(f"{reverse('kitchen:location_based_prices')}?location_id={location_id}")


# ==========================================================================
# BEIT NAHLA KITCHEN PANEL
# ==========================================================================
# Single-page panel with AJAX-only CRUD so changes don't reload the screen.
# All write endpoints return JSON; the template manipulates DOM in place.

from catering.models import (
    BeitNahlaSettings, BeitNahlaDistanceTier,
    BeitNahlaMealBox, BeitNahlaMealBoxImage,
    BeitNahlaOptionCategory, BeitNahlaOptionItem,
    BeitNahlaOrder, BeitNahlaOrderItem, BeitNahlaOrderStatus,
)
from django.views.decorators.http import require_http_methods


def _bn_meal_box_payload(box, request):
    images = [
        {
            "id": img.id,
            "url": request.build_absolute_uri(img.image.url) if img.image else None,
            "order": img.order,
        }
        for img in box.images.all().order_by("order", "id")
    ]
    return {
        "id": box.id,
        "name": box.name,
        "description": box.description or "",
        "image_url": request.build_absolute_uri(box.image.url) if box.image else None,
        "images": images,
        "display_order": box.display_order,
        "is_active": box.is_active,
    }


def _bn_category_payload(cat):
    return {
        "id": cat.id,
        "name": cat.name,
        "description": cat.description or "",
        "display_order": cat.display_order,
        "is_active": cat.is_active,
        "items_count": cat.items.count(),
    }


def _bn_item_payload(item, request):
    return {
        "id": item.id,
        "category_id": item.category_id,
        "name": item.name,
        "description": item.description or "",
        "image_url": request.build_absolute_uri(item.image.url) if item.image else None,
        "display_order": item.display_order,
        "is_active": item.is_active,
    }


def _bn_tier_payload(tier):
    return {
        "id": tier.id,
        "label": tier.label or "",
        "min_km": str(tier.min_km),
        "max_km": str(tier.max_km),
        "service_charge": str(tier.service_charge),
        "delivery_charge": str(tier.delivery_charge),
        "is_active": tier.is_active,
    }


def _bn_settings_payload(cfg):
    return {
        "order_now_price": str(cfg.order_now_price),
        "weekly_price": str(cfg.weekly_price),
        "restaurant_name": cfg.restaurant_name,
        "restaurant_latitude": str(cfg.restaurant_latitude),
        "restaurant_longitude": str(cfg.restaurant_longitude),
        "max_deliverable_km": str(cfg.max_deliverable_km),
        "opening_time": cfg.opening_time.strftime("%H:%M"),
        "closing_time": cfg.closing_time.strftime("%H:%M"),
    }


@login_required
@user_passes_test(is_kitchen_admin)
def beit_nahla_panel_view(request):
    """Render the single Beit Nahla admin page (loads initial JSON inline)."""
    cfg = BeitNahlaSettings.load()
    boxes = BeitNahlaMealBox.objects.prefetch_related("images").order_by(
        "display_order", "name"
    )
    categories = BeitNahlaOptionCategory.objects.prefetch_related("items").order_by(
        "display_order", "name"
    )
    tiers = BeitNahlaDistanceTier.objects.order_by("min_km", "id")

    context = {
        "settings_json": _bn_settings_payload(cfg),
        "boxes_json": [_bn_meal_box_payload(b, request) for b in boxes],
        "categories_json": [
            {
                **_bn_category_payload(c),
                "items": [_bn_item_payload(i, request) for i in c.items.all().order_by(
                    "display_order", "name"
                )],
            }
            for c in categories
        ],
        "tiers_json": [_bn_tier_payload(t) for t in tiers],
    }
    return render(request, "kitchen/beit_nahla.html", context)


# ---------- Settings ----------

@login_required
@user_passes_test(is_kitchen_admin)
@require_http_methods(["POST"])
def beit_nahla_settings_api(request):
    cfg = BeitNahlaSettings.load()
    fields = [
        "order_now_price", "weekly_price",
        "restaurant_name", "restaurant_latitude", "restaurant_longitude",
        "max_deliverable_km", "opening_time", "closing_time",
    ]
    for f in fields:
        if f in request.POST:
            val = request.POST[f].strip()
            if val == "":
                continue
            setattr(cfg, f, val)
    try:
        cfg.full_clean()
        cfg.save()
    except Exception as e:
        return JsonResponse({"ok": False, "error": str(e)}, status=400)
    return JsonResponse({"ok": True, "settings": _bn_settings_payload(cfg)})


# ---------- Distance tiers ----------

@login_required
@user_passes_test(is_kitchen_admin)
@require_http_methods(["POST"])
def beit_nahla_tier_create_api(request):
    try:
        tier = BeitNahlaDistanceTier.objects.create(
            label=request.POST.get("label", "").strip(),
            min_km=request.POST.get("min_km") or 0,
            max_km=request.POST.get("max_km") or 0,
            service_charge=request.POST.get("service_charge") or 0,
            delivery_charge=request.POST.get("delivery_charge") or 0,
            is_active=request.POST.get("is_active", "true").lower() == "true",
        )
    except Exception as e:
        return JsonResponse({"ok": False, "error": str(e)}, status=400)
    return JsonResponse({"ok": True, "tier": _bn_tier_payload(tier)})


@login_required
@user_passes_test(is_kitchen_admin)
@require_http_methods(["POST"])
def beit_nahla_tier_update_api(request, pk):
    tier = get_object_or_404(BeitNahlaDistanceTier, pk=pk)
    for f in ("label", "min_km", "max_km", "service_charge", "delivery_charge"):
        if f in request.POST:
            v = request.POST[f].strip()
            if f == "label":
                setattr(tier, f, v)
            elif v != "":
                setattr(tier, f, v)
    if "is_active" in request.POST:
        tier.is_active = request.POST["is_active"].lower() == "true"
    try:
        tier.save()
    except Exception as e:
        return JsonResponse({"ok": False, "error": str(e)}, status=400)
    return JsonResponse({"ok": True, "tier": _bn_tier_payload(tier)})


@login_required
@user_passes_test(is_kitchen_admin)
@require_http_methods(["POST"])
def beit_nahla_tier_delete_api(request, pk):
    tier = get_object_or_404(BeitNahlaDistanceTier, pk=pk)
    tier.delete()
    return JsonResponse({"ok": True})


# ---------- Meal boxes ----------

@login_required
@user_passes_test(is_kitchen_admin)
@require_http_methods(["POST"])
def beit_nahla_meal_box_create_api(request):
    try:
        box = BeitNahlaMealBox.objects.create(
            name=request.POST.get("name", "").strip(),
            description=request.POST.get("description", "").strip(),
            display_order=int(request.POST.get("display_order") or 0),
            is_active=request.POST.get("is_active", "true").lower() == "true",
        )
        if "image" in request.FILES:
            box.image = request.FILES["image"]
            box.save()
    except Exception as e:
        return JsonResponse({"ok": False, "error": str(e)}, status=400)
    return JsonResponse({"ok": True, "box": _bn_meal_box_payload(box, request)})


@login_required
@user_passes_test(is_kitchen_admin)
@require_http_methods(["POST"])
def beit_nahla_meal_box_update_api(request, pk):
    box = get_object_or_404(BeitNahlaMealBox, pk=pk)
    if "name" in request.POST:
        box.name = request.POST["name"].strip()
    if "description" in request.POST:
        box.description = request.POST["description"].strip()
    if "display_order" in request.POST:
        try:
            box.display_order = int(request.POST["display_order"] or 0)
        except (TypeError, ValueError):
            pass
    if "is_active" in request.POST:
        box.is_active = request.POST["is_active"].lower() == "true"
    if "image" in request.FILES:
        box.image = request.FILES["image"]
    try:
        box.save()
    except Exception as e:
        return JsonResponse({"ok": False, "error": str(e)}, status=400)
    return JsonResponse({"ok": True, "box": _bn_meal_box_payload(box, request)})


@login_required
@user_passes_test(is_kitchen_admin)
@require_http_methods(["POST"])
def beit_nahla_meal_box_delete_api(request, pk):
    box = get_object_or_404(BeitNahlaMealBox, pk=pk)
    box.delete()
    return JsonResponse({"ok": True})


# ---------- Option categories ----------

@login_required
@user_passes_test(is_kitchen_admin)
@require_http_methods(["POST"])
def beit_nahla_category_create_api(request):
    try:
        cat = BeitNahlaOptionCategory.objects.create(
            name=request.POST.get("name", "").strip(),
            description=request.POST.get("description", "").strip(),
            display_order=int(request.POST.get("display_order") or 0),
            is_active=request.POST.get("is_active", "true").lower() == "true",
        )
    except Exception as e:
        return JsonResponse({"ok": False, "error": str(e)}, status=400)
    payload = _bn_category_payload(cat)
    payload["items"] = []
    return JsonResponse({"ok": True, "category": payload})


@login_required
@user_passes_test(is_kitchen_admin)
@require_http_methods(["POST"])
def beit_nahla_category_update_api(request, pk):
    cat = get_object_or_404(BeitNahlaOptionCategory, pk=pk)
    if "name" in request.POST:
        cat.name = request.POST["name"].strip()
    if "description" in request.POST:
        cat.description = request.POST["description"].strip()
    if "display_order" in request.POST:
        try:
            cat.display_order = int(request.POST["display_order"] or 0)
        except (TypeError, ValueError):
            pass
    if "is_active" in request.POST:
        cat.is_active = request.POST["is_active"].lower() == "true"
    try:
        cat.save()
    except Exception as e:
        return JsonResponse({"ok": False, "error": str(e)}, status=400)
    return JsonResponse({"ok": True, "category": _bn_category_payload(cat)})


@login_required
@user_passes_test(is_kitchen_admin)
@require_http_methods(["POST"])
def beit_nahla_category_delete_api(request, pk):
    cat = get_object_or_404(BeitNahlaOptionCategory, pk=pk)
    cat.delete()
    return JsonResponse({"ok": True})


# ---------- Option items ----------

@login_required
@user_passes_test(is_kitchen_admin)
@require_http_methods(["POST"])
def beit_nahla_item_create_api(request):
    cat_id = request.POST.get("category_id")
    cat = get_object_or_404(BeitNahlaOptionCategory, pk=cat_id)
    try:
        item = BeitNahlaOptionItem.objects.create(
            category=cat,
            name=request.POST.get("name", "").strip(),
            description=request.POST.get("description", "").strip(),
            display_order=int(request.POST.get("display_order") or 0),
            is_active=request.POST.get("is_active", "true").lower() == "true",
        )
        if "image" in request.FILES:
            item.image = request.FILES["image"]
            item.save()
    except Exception as e:
        return JsonResponse({"ok": False, "error": str(e)}, status=400)
    return JsonResponse({"ok": True, "item": _bn_item_payload(item, request)})


@login_required
@user_passes_test(is_kitchen_admin)
@require_http_methods(["POST"])
def beit_nahla_item_update_api(request, pk):
    item = get_object_or_404(BeitNahlaOptionItem, pk=pk)
    if "name" in request.POST:
        item.name = request.POST["name"].strip()
    if "description" in request.POST:
        item.description = request.POST["description"].strip()
    if "display_order" in request.POST:
        try:
            item.display_order = int(request.POST["display_order"] or 0)
        except (TypeError, ValueError):
            pass
    if "is_active" in request.POST:
        item.is_active = request.POST["is_active"].lower() == "true"
    if "image" in request.FILES:
        item.image = request.FILES["image"]
    try:
        item.save()
    except Exception as e:
        return JsonResponse({"ok": False, "error": str(e)}, status=400)
    return JsonResponse({"ok": True, "item": _bn_item_payload(item, request)})


@login_required
@user_passes_test(is_kitchen_admin)
@require_http_methods(["POST"])
def beit_nahla_item_delete_api(request, pk):
    item = get_object_or_404(BeitNahlaOptionItem, pk=pk)
    item.delete()
    return JsonResponse({"ok": True})


# ---------- Beit Nahla orders (read + status updates) ----------

def _bn_order_payload(o):
    return {
        "id": o.id,
        "order_id": o.order_id,
        "status": o.status,
        "status_display": o.get_status_display(),
        "mode": o.mode,
        "customer_name": o.customer_name,
        "customer_phone": o.customer_phone,
        "building": o.building,
        "street": o.street,
        "appt": o.appt,
        "delivery_address": o.delivery_address,
        "latitude": str(o.latitude) if o.latitude is not None else None,
        "longitude": str(o.longitude) if o.longitude is not None else None,
        "distance_km": str(o.distance_km) if o.distance_km is not None else None,
        "tier_label": o.tier_label,
        "subtotal": str(o.subtotal),
        "vat": str(o.vat),
        "discount": str(o.discount),
        "service_charge": str(o.service_charge),
        "delivery_charge": str(o.delivery_charge),
        "total_amount": str(o.total_amount),
        "user": o.user.username if o.user else None,
        "created_at": o.created_at.isoformat(),
        "items": [
            {
                "id": it.id,
                "box_name": it.box_name,
                "unit_price": str(it.unit_price),
                "quantity": it.quantity,
                "selections_summary": it.selections_summary,
            }
            for it in o.items.all()
        ],
    }


@login_required
@user_passes_test(is_kitchen_admin)
def beit_nahla_orders_view(request):
    orders = (
        BeitNahlaOrder.objects
        .prefetch_related('items')
        .order_by('-created_at')[:200]
    )
    return render(request, 'kitchen/beit_nahla_orders.html', {
        'orders_json': [_bn_order_payload(o) for o in orders],
        'status_choices': [
            {'value': v, 'label': l} for v, l in BeitNahlaOrderStatus.choices
        ],
    })


@login_required
@user_passes_test(is_kitchen_admin)
@require_http_methods(["POST"])
def beit_nahla_order_status_api(request, pk):
    order = get_object_or_404(BeitNahlaOrder, pk=pk)
    new_status = request.POST.get('status')
    valid = {v for v, _ in BeitNahlaOrderStatus.choices}
    if new_status not in valid:
        return JsonResponse({"ok": False, "error": "Invalid status"}, status=400)
    order.status = new_status
    order.save(update_fields=['status', 'updated_at'])
    return JsonResponse({"ok": True, "order": _bn_order_payload(order)})


@login_required
@user_passes_test(is_kitchen_admin)
def get_active_beit_nahla_orders_api(request):
    """
    Returns JSON list of latest Beit Nahla orders for polling.
    """
    orders = (
        BeitNahlaOrder.objects
        .prefetch_related('items')
        .order_by('-created_at')[:200]
    )
    return JsonResponse({
        "ok": True,
        "orders": [_bn_order_payload(o) for o in orders]
    })
